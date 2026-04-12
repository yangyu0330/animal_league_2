from __future__ import annotations

import re
import sys
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPREADSHEET_PATH = ROOT / "tmp" / "spreadsheets" / "academyinfo_departments_20241007.xlsx"
OUTPUT_PATH = ROOT / "docs" / "week2" / "B-5_seed_likelion_university_data.sql"
ACADEMYINFO_DETAIL_URL = (
    "https://www.academyinfo.go.kr/brd/brd0520/selectDetail.do?"
    "ntce_sntc_sno=160&bbs_gubun=rfbr&no=19"
)
ACADEMYINFO_DOWNLOAD_URL = "https://www.academyinfo.go.kr/file/FileDown.do"
ACADEMYINFO_FILE_NO = "4580256"


SCHOOLS = [
    {"name": "가천대", "region": "경기/인천", "official": ["가천대학교"]},
    {"name": "가톨릭대", "region": "경기/인천", "official": ["가톨릭대학교"]},
    {"name": "강남대", "region": "경기/인천", "official": ["강남대학교"]},
    {"name": "계명대", "region": "경상", "official": ["계명대학교"]},
    {"name": "고려대(세종)", "region": "충청", "official": ["고려대학교(세종)"]},
    {"name": "고려대(서울)", "region": "서울", "official": ["고려대학교"]},
    {"name": "경기대", "region": "경기/인천", "official": ["경기대학교"], "locations": ["경기"]},
    {"name": "경상국립대", "region": "경상", "official": ["경상국립대학교"]},
    {"name": "경성대", "region": "경상", "official": ["경성대학교"]},
    {"name": "공주대(천안)", "region": "충청", "official": ["국립공주대학교"], "locations": ["충남", "천안"]},
    {"name": "경북대", "region": "경상", "official": ["경북대학교"]},
    {"name": "경희대", "region": "서울/경기", "official": ["경희대학교"]},
    {"name": "광운대", "region": "서울", "official": ["광운대학교"]},
    {"name": "국민대", "region": "서울", "official": ["국민대학교"]},
    {"name": "금오공대", "region": "경상", "official": ["국립금오공과대학교"]},
    {"name": "순천대", "region": "전라", "official": ["국립순천대학교"]},
    {"name": "단국대", "region": "경기/인천", "official": ["단국대학교"], "locations": ["경기", "죽전"]},
    {"name": "국립창원대", "region": "경상", "official": ["국립창원대학교", "창원대학교"]},
    {"name": "덕성여대", "region": "서울", "official": ["덕성여자대학교"]},
    {"name": "동국대", "region": "서울", "official": ["동국대학교"]},
    {"name": "동덕여대", "region": "서울", "official": ["동덕여자대학교"]},
    {"name": "명지대(인문)", "region": "서울", "official": ["명지대학교"], "locations": ["서울"]},
    {"name": "대구대", "region": "경상", "official": ["대구대학교"]},
    {"name": "동아대", "region": "경상", "official": ["동아대학교"]},
    {"name": "백석대", "region": "충청", "official": ["백석대학교"]},
    {"name": "삼육대", "region": "서울", "official": ["삼육대학교"]},
    {"name": "상명대(천안)", "region": "충청", "official": ["상명대학교"], "locations": ["충남", "천안"]},
    {"name": "상명대(서울)", "region": "서울", "official": ["상명대학교"], "locations": ["서울"]},
    {"name": "서강대", "region": "서울", "official": ["서강대학교"]},
    {"name": "명지대(자연)", "region": "경기/인천", "official": ["명지대학교"], "locations": ["경기", "용인"]},
    {"name": "서경대", "region": "서울", "official": ["서경대학교"]},
    {"name": "서울과기대", "region": "서울", "official": ["서울과학기술대학교"]},
    {"name": "서울대", "region": "서울", "official": ["서울대학교"]},
    {"name": "서울여대", "region": "서울", "official": ["서울여자대학교"]},
    {"name": "서울시립대", "region": "서울", "official": ["서울시립대학교"]},
    {"name": "서울예대", "region": "서울", "official": ["서울예술대학교"]},
    {"name": "성결대", "region": "경기/인천", "official": ["성결대학교"]},
    {"name": "성공회대", "region": "서울", "official": ["성공회대학교"]},
    {"name": "성균관대", "region": "서울", "official": ["성균관대학교"]},
    {"name": "성신여대", "region": "서울", "official": ["성신여자대학교"]},
    {"name": "서일대", "region": "서울", "official": ["서일대학교"]},
    {"name": "선문대", "region": "충청", "official": ["선문대학교"]},
    {"name": "숙명여대", "region": "서울", "official": ["숙명여자대학교"]},
    {"name": "순천향대", "region": "충청", "official": ["순천향대학교"]},
    {"name": "숭실대", "region": "서울", "official": ["숭실대학교"]},
    {"name": "연세대", "region": "서울", "official": ["연세대학교"]},
    {"name": "영남대", "region": "경상", "official": ["영남대학교"]},
    {"name": "을지대", "region": "충청", "official": ["을지대학교"]},
    {"name": "이화여대", "region": "서울", "official": ["이화여자대학교"]},
    {"name": "연세대(미래)", "region": "강원", "official": ["연세대학교 미래캠퍼스", "연세대학교(미래)"]},
    {"name": "수원대", "region": "경기/인천", "official": ["수원대학교"]},
    {"name": "신한대", "region": "경기/인천", "official": ["신한대학교"]},
    {"name": "연암공대", "region": "충청", "official": ["연암공과대학교"]},
    {"name": "우송대", "region": "충청", "official": ["우송대학교"]},
    {"name": "인천대", "region": "경기/인천", "official": ["인천대학교"]},
    {"name": "인하대", "region": "경기/인천", "official": ["인하대학교"]},
    {"name": "중부대", "region": "경기/인천", "official": ["중부대학교"], "locations": ["경기", "고양"]},
    {"name": "중앙대", "region": "서울", "official": ["중앙대학교"]},
    {"name": "청주대", "region": "충청", "official": ["청주대학교"]},
    {"name": "충남대", "region": "충청", "official": ["충남대학교"]},
    {"name": "한국교통대", "region": "충청", "official": ["국립한국교통대학교"]},
    {"name": "한국외대(글로벌)", "region": "경기/인천", "official": ["한국외국어대학교"], "locations": ["경기", "용인"]},
    {"name": "한국외대", "region": "서울", "official": ["한국외국어대학교"], "locations": ["서울"]},
    {"name": "한국항공대", "region": "경기/인천", "official": ["한국항공대학교"]},
    {"name": "한남대", "region": "충청", "official": ["한남대학교"]},
    {"name": "한동대", "region": "경상", "official": ["한동대학교"]},
    {"name": "한밭대", "region": "충청", "official": ["국립한밭대학교"]},
    {"name": "한서대", "region": "충청", "official": ["한서대학교"]},
    {"name": "한성대", "region": "서울", "official": ["한성대학교"]},
    {"name": "한림대", "region": "강원", "official": ["한림대학교"]},
    {"name": "제주대", "region": "제주", "official": ["제주대학교"]},
    {"name": "인하공전", "region": "경기/인천", "official": ["인하공업전문대학"]},
    {"name": "태재대", "region": "경기/인천", "official": ["태재대학교"]},
    {"name": "평택대", "region": "경기/인천", "official": ["평택대학교"]},
    {"name": "한국공학대", "region": "경기/인천", "official": ["한국공학대학교"]},
    {"name": "전북대", "region": "전라", "official": ["전북대학교"]},
    {"name": "한국기술교육대", "region": "충청", "official": ["한국기술교육대학교"]},
    {"name": "충북대", "region": "충청", "official": ["충북대학교"]},
    {"name": "홍익대", "region": "서울", "official": ["홍익대학교"]},
    {"name": "호서대", "region": "충청", "official": ["호서대학교"]},
]


def ensure_spreadsheet() -> None:
    if SPREADSHEET_PATH.exists():
        return

    SPREADSHEET_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = urllib.parse.urlencode({"atch_file_no": ACADEMYINFO_FILE_NO}).encode()
    request = urllib.request.Request(
        ACADEMYINFO_DOWNLOAD_URL,
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        SPREADSHEET_PATH.write_bytes(response.read())


def sql(value: object) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def normalize_department_name(name: str) -> str:
    return re.sub(r"[\s\-_.()/]", "", name).lower()


def map_category(major_class: str, mid_class: str, department_name: str) -> str:
    field_text = f"{mid_class} {department_name}"
    all_text = f"{major_class} {field_text}"
    if "교육" in field_text or department_name.endswith("교육과"):
        return "교육"
    if any(
        token in all_text
        for token in (
            "의약",
            "보건",
            "간호",
            "의학",
            "의예",
            "약학",
            "제약",
            "치위생",
            "물리치료",
            "작업치료",
        )
    ):
        return "보건/의학"
    if any(token in field_text for token in ("경영", "경제", "회계", "세무", "무역", "금융")):
        return "경영/경제"
    if "공학" in all_text:
        return "공학"
    if "자연" in all_text:
        return "자연과학"
    if any(token in all_text for token in ("예체능", "예술", "체육", "디자인", "음악", "미술")):
        return "예술/체육"
    if any(token in field_text for token in ("언어", "문학", "인문", "역사", "철학", "국어", "영어", "중어", "일어")):
        return "인문"
    return "사회과학"


def template_for(category: str) -> str:
    return {
        "공학": "eng_default_01",
        "자연과학": "science_default_01",
        "인문": "humanities_default_01",
        "사회과학": "social_default_01",
        "경영/경제": "biz_default_01",
        "예술/체육": "arts_default_01",
        "교육": "edu_default_01",
        "보건/의학": "health_default_01",
    }[category]


def location_matches(row: tuple[object, ...], locations: list[str] | None) -> bool:
    if not locations:
        return True
    fields = " ".join(str(row[i] or "") for i in (7, 21, 22))
    return any(location in fields for location in locations)


def load_department_rows() -> tuple[list[dict[str, object]], Counter[str]]:
    ensure_spreadsheet()
    wb = load_workbook(SPREADSHEET_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Row 5 columns:
    # 2 school, 7 university region, 11 department, 14 status,
    # 15 major class, 16 middle class, 20 degree course, 21/22 location.
    source_rows = list(ws.iter_rows(min_row=6, values_only=True))
    departments: list[dict[str, object]] = []
    school_counts: Counter[str] = Counter()
    seen_by_school: dict[str, set[str]] = {}

    for school in SCHOOLS:
        seen_by_school[school["name"]] = set()
        official_names = set(school["official"])
        for row in source_rows:
            official_school_name = str(row[2] or "")
            if official_school_name not in official_names:
                continue
            if not location_matches(row, school.get("locations")):
                continue

            status = str(row[14] or "")
            if "폐지" in status:
                continue
            degree_course = str(row[20] or "")
            if degree_course not in {"학사", "전문학사"}:
                continue

            department_name = str(row[11] or "").strip()
            if not department_name:
                continue

            normalized_name = normalize_department_name(department_name)
            if normalized_name in seen_by_school[school["name"]]:
                continue
            seen_by_school[school["name"]].add(normalized_name)

            category = map_category(str(row[15] or ""), str(row[16] or ""), department_name)
            departments.append(
                {
                    "school_name": school["name"],
                    "department_name": department_name,
                    "normalized_name": normalized_name,
                    "category": category,
                    "template_id": template_for(category),
                    "official_school_name": official_school_name,
                    "source_region": str(row[7] or ""),
                    "department_location": str(row[21] or ""),
                    "source_status": status,
                }
            )
            school_counts[school["name"]] += 1

    return departments, school_counts


def render_sql(departments: list[dict[str, object]], school_counts: Counter[str]) -> str:
    school_values = ",\n".join(
        f"    ({sql(school['name'])}, {sql(school['region'])})"
        for school in SCHOOLS
    )
    department_values = ",\n".join(
        "    ("
        + ", ".join(
            [
                sql(row["school_name"]),
                sql(row["department_name"]),
                sql(row["normalized_name"]),
                sql(row["category"]),
                sql(row["template_id"]),
                sql(row["official_school_name"]),
                sql(row["source_region"]),
                sql(row["department_location"]),
                sql(row["source_status"]),
            ]
        )
        + ")"
        for row in departments
    )
    missing = [school["name"] for school in SCHOOLS if school_counts[school["name"]] == 0]
    summary = "\n".join(
        f"-- {school['name']}: {school_counts[school['name']]} departments"
        for school in SCHOOLS
    )

    return f"""-- B-5: Replace dummy seeded data with 14th LikeLion university schools and official department data.
-- Generated by scripts/generate_likelion_seed.py.
-- Sources:
-- - LikeLion university 14th active schools: https://likelion.university/
-- - AcademyInfo department workbook: {ACADEMYINFO_DETAIL_URL}
-- - AcademyInfo workbook basis date: 2024-10-07
--
-- Notes:
-- - AcademyInfo states this workbook is for public disclosure/statistical education-unit management.
-- - Rows whose department status contains '폐지' are excluded.
-- - Existing rows under school.is_seeded=true are cleared before this seed is inserted.
-- - At least one app_user row is required because department.created_by is not nullable.
-- - User-provided '연남대' is normalized to the official LikeLion/AcademyInfo match '영남대'.
-- - Missing matches: {', '.join(missing) if missing else 'none'}
-- - Supabase SQL Editor may show "Failed to generate title" for this large file.
--   That toast is from the editor title helper, not from Postgres execution.
--
-- Summary:
-- Schools: {len(SCHOOLS)}
-- Departments: {len(departments)}
{summary}

begin;

do $$
declare
  seed_owner uuid;
begin
  select id
    into seed_owner
  from app_user
  order by created_at asc
  limit 1;

  if seed_owner is null then
    raise exception 'No rows found in app_user. Sign in once before running this seed.';
  end if;
end $$;

with seeded_departments as (
  select d.id
  from department d
  join school s
    on s.id = d.school_id
  where s.is_seeded = true
),
cleared_selected_departments as (
  update app_user
  set
    selected_department_id = null,
    updated_at = now()
  where selected_department_id in (select id from seeded_departments)
  returning id
),
seeded_schools as (
  select id
  from school
  where is_seeded = true
),
cleared_selected_schools as (
  update app_user
  set
    selected_school_id = null,
    selected_department_id = null,
    updated_at = now()
  where selected_school_id in (select id from seeded_schools)
  returning id
),
deleted_departments as (
  delete from department
  where id in (select id from seeded_departments)
  returning id
)
delete from school
where is_seeded = true;

with seed_schools (name, region) as (
  values
{school_values}
)
insert into school (name, region, is_seeded)
select name, region, true
from seed_schools
on conflict (name) do update
set
  region = excluded.region,
  is_seeded = true;

with
seed_owner as (
  select id
  from app_user
  order by created_at asc
  limit 1
),
seed_departments (
  school_name,
  department_name,
  normalized_name,
  category,
  template_id,
  official_school_name,
  source_region,
  department_location,
  source_status
) as (
  values
{department_values}
)
insert into department (
  school_id,
  name,
  normalized_name,
  category,
  template_id,
  total_clicks,
  accepted_clicks,
  pressure_level,
  created_by
)
select
  s.id,
  d.department_name,
  d.normalized_name,
  d.category,
  d.template_id,
  0,
  0,
  0,
  o.id
from seed_departments d
join school s
  on s.name = d.school_name
cross join seed_owner o
on conflict (school_id, normalized_name) do update
set
  name = excluded.name,
  category = excluded.category,
  template_id = excluded.template_id;

commit;

-- Verification queries after running:
-- select count(*) from school where is_seeded = true;
-- select count(*) from department d join school s on s.id = d.school_id where s.is_seeded = true;
"""


def main() -> int:
    departments, school_counts = load_department_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(render_sql(departments, school_counts), encoding="utf-8", newline="\n")

    print(f"schools={len(SCHOOLS)} departments={len(departments)}")
    for school in SCHOOLS:
        print(f"{school['name']}: {school_counts[school['name']]}")
    missing = [school["name"] for school in SCHOOLS if school_counts[school["name"]] == 0]
    if missing:
        print("missing=" + ", ".join(missing), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
